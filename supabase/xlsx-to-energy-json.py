#!/usr/bin/env python3
"""Converts the research workbook into src/data/energy.content.json.

Usage (from the website/ dir):
    python supabase/xlsx-to-energy-json.py ../us_renewable_policy_data.xlsx

The workbook (5 sheets: Overview, Policies, Documents, Policy_Document_Links,
Status_Assessments) is the upstream research artifact; the generated JSON is
what the app and the DB seed actually read. Re-run this whenever a refreshed
workbook lands, then re-run `node supabase/generate-energy-seed.mjs`.

Only the value shapes change here: dates are normalised to ISO `YYYY-MM-DD`,
pipe-delimited cells become arrays, and Excel booleans become real booleans.
No field is dropped, so the JSON stays a faithful copy of the workbook.
"""
import json
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - dependency hint
    sys.exit("openpyxl is required:  pip install openpyxl")

# Cells that hold a "a | b | c" list rather than a single value.
LIST_FIELDS = {
    "lead_agencies",
    "affected_entities",
    "supporting_document_ids",
}

# Sheet name -> key in the emitted JSON.
SHEETS = {
    "Policies": "policies",
    "Documents": "documents",
    "Policy_Document_Links": "links",
    "Status_Assessments": "assessments",
}


def normalize(key, value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return None
    if key in LIST_FIELDS:
        return [part.strip() for part in text.split("|") if part.strip()]
    # Excel stores the TRUE/FALSE flags as text in some exports.
    if text in ("True", "TRUE"):
        return True
    if text in ("False", "FALSE"):
        return False
    return text


def read_sheet(worksheet):
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        for header, cell in zip(headers, row):
            if not header:
                continue
            value = normalize(header, cell)
            if value is not None:
                record[header] = value
        if record:
            records.append(record)
    return records


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    workbook_path = Path(sys.argv[1]).resolve()
    out_path = Path(__file__).resolve().parent.parent / "src" / "data" / "energy.content.json"

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    payload = {
        "meta": {
            "title": "U.S. Renewable Energy Policy — Current Status",
            "source": workbook_path.name,
            "generatedFrom": "supabase/xlsx-to-energy-json.py",
        },
    }
    for sheet_name, json_key in SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            sys.exit(f"Workbook is missing the {sheet_name!r} sheet")
        payload[json_key] = read_sheet(workbook[sheet_name])

    # `status_as_of` is uniform across the workbook; surface it as the dataset
    # date so the UI can state how current the content is.
    dates = {p.get("status_as_of") for p in payload["policies"] if p.get("status_as_of")}
    payload["meta"]["statusAsOf"] = sorted(dates)[-1] if dates else None

    out_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    counts = ", ".join(f"{len(payload[k])} {k}" for k in SHEETS.values())
    print(f"Wrote {out_path} ({counts})")


if __name__ == "__main__":
    main()
