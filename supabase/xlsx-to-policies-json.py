#!/usr/bin/env python3
"""Convert the renewable-policy workbook into canonical policy JSON.

Usage (from the website/ dir):
    python supabase/xlsx-to-policies-json.py ../us_renewable_policy_data.xlsx
"""
import json
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - dependency hint
    sys.exit("openpyxl is required:  pip install openpyxl")


LIST_FIELDS = {
    "lead_agencies",
    "affected_entities",
    "supporting_document_ids",
}

SHEETS = {
    "Policies": "policies",
    "Documents": "documents",
    "Policy_Document_Links": "links",
    "Status_Assessments": "assessments",
}


def normalize(key, value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return None
    if key in LIST_FIELDS:
        return [part.strip() for part in text.split("|") if part.strip()]
    if text in ("True", "TRUE"):
        return True
    if text in ("False", "FALSE"):
        return False
    return text


def read_sheet(worksheet):
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(header).strip() if header is not None else "" for header in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        for header, cell in zip(headers, row):
            if not header:
                continue
            value = normalize(header, cell)
            if value is not None:
                record[header] = value
        if record:
            records.append(record)
    return records


def area_lookup(area_config):
    lookup = {}
    for area in area_config["areas"]:
        for source_value in area["source_values"]:
            if source_value in lookup:
                raise ValueError(f"Duplicate source policy area mapping: {source_value}")
            lookup[source_value] = area["slug"]
    return lookup


def classify_policies(policies, lookup):
    for index, policy in enumerate(policies):
        policy_id = policy.get("policy_id", f"row {index + 2}")
        source_value = policy.pop("policy_area", None)
        if source_value not in lookup:
            raise ValueError(f"Unknown policy area for {policy_id}: {source_value}")
        policy["source_policy_area"] = source_value
        policy["policy_area_slug"] = lookup[source_value]
        policy["sort_order"] = index


def unique_ids(rows, key):
    ids = set()
    for row in rows:
        identifier = row.get(key)
        if identifier in ids:
            raise ValueError(f"Duplicate {key}: {identifier}")
        ids.add(identifier)
    return ids


def validate_dataset(payload):
    policy_ids = unique_ids(payload["policies"], "policy_id")
    document_ids = unique_ids(payload["documents"], "document_id")
    unique_ids(payload["links"], "link_id")
    unique_ids(payload["assessments"], "assessment_id")

    for link in payload["links"]:
        if link.get("policy_id") not in policy_ids:
            raise ValueError(
                f"{link.get('link_id')} references unknown policy {link.get('policy_id')}"
            )
        if link.get("document_id") not in document_ids:
            raise ValueError(
                f"{link.get('link_id')} references unknown document {link.get('document_id')}"
            )

    assessments_by_policy = {policy_id: set() for policy_id in policy_ids}
    for assessment in payload["assessments"]:
        assessment_id = assessment.get("assessment_id")
        policy_id = assessment.get("policy_id")
        if policy_id not in policy_ids:
            raise ValueError(f"{assessment_id} references unknown policy {policy_id}")
        assessments_by_policy[policy_id].add(assessment.get("assessment_type"))

        primary_document_id = assessment.get("primary_document_id")
        if primary_document_id is not None and primary_document_id not in document_ids:
            raise ValueError(
                f"{assessment_id} references unknown primary document {primary_document_id}"
            )
        for supporting_document_id in assessment.get("supporting_document_ids", []):
            if supporting_document_id not in document_ids:
                raise ValueError(
                    f"{assessment_id} references unknown supporting document {supporting_document_id}"
                )

    required_assessments = (
        "legal_status",
        "implementation_status",
        "litigation_status",
    )
    for policy in payload["policies"]:
        policy_id = policy.get("policy_id")
        present = assessments_by_policy.get(policy_id, set())
        for assessment_type in required_assessments:
            if assessment_type not in present:
                raise ValueError(f"{policy_id} is missing {assessment_type}")


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)

    workbook_path = Path(sys.argv[1]).resolve()
    data_dir = Path(__file__).resolve().parent.parent / "src" / "data"
    out_path = data_dir / "policies.content.json"
    tmp_path = Path(f"{out_path}.tmp")
    area_config = json.loads(
        (data_dir / "policy-areas.content.json").read_text(encoding="utf-8")
    )

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    try:
        payload = {
            "meta": {
                "title": "U.S. Renewable Energy Policy — Current Status",
                "source": workbook_path.name,
                "generatedFrom": "supabase/xlsx-to-policies-json.py",
            },
        }
        for sheet_name, json_key in SHEETS.items():
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Workbook is missing the {sheet_name!r} sheet")
            payload[json_key] = read_sheet(workbook[sheet_name])
    finally:
        workbook.close()

    classify_policies(payload["policies"], area_lookup(area_config))
    dates = {
        policy.get("status_as_of")
        for policy in payload["policies"]
        if policy.get("status_as_of")
    }
    payload["meta"]["statusAsOf"] = sorted(dates)[-1] if dates else None
    validate_dataset(payload)

    tmp_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    tmp_path.replace(out_path)
    counts = ", ".join(f"{len(payload[key])} {key}" for key in SHEETS.values())
    print(f"Wrote {out_path} ({counts})")


if __name__ == "__main__":
    main()
